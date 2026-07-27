#!/usr/bin/env python
"""Batch extract first-pass RoHS data from PDF test reports.

Put PDFs in input_pdfs, then run:
    python rohs_pdf_extract.py

Outputs stay inside this project folder:
    intermediate_json/*.json
    outputs/rohs_result.xlsx
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import shutil
import sys
import urllib.error
import urllib.request
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


logging.getLogger("pypdf").setLevel(logging.ERROR)

ROOT = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
DEFAULT_INPUT = ROOT / "input_pdfs"
DEFAULT_JSON = ROOT / "intermediate_json"
DEFAULT_OUTPUT = ROOT / "outputs" / "rohs_result.xlsx"
DEFAULT_AI_LOG = ROOT / "logs" / "ai_assist_log.jsonl"
DEFAULT_AI_LOG_FILES = ROOT / "logs" / "ai_assist_files"


SUBSTANCES = [
    ("Lead", ["lead", "pb", "铅", "鉛"]),
    ("Mercury", ["mercury", "hg", "汞"]),
    ("Cadmium", ["cadmium", "cd", "镉", "鎘"]),
    (
        "Hexavalent Chromium",
        ["hexavalent chromium", "chromium vi", "cr(vi)", "cr6+", "六价铬", "六價鉻"],
    ),
    ("PBBs", ["pbb", "pbbs", "polybrominated biphenyl", "多溴联苯", "多溴聯苯"]),
    ("PBDEs", ["pbde", "pbdes", "polybrominated diphenyl ether", "多溴二苯醚"]),
    ("BBP", ["bbp", "benzyl butyl phthalate", "butyl benzyl phthalate", "邻苯二甲酸丁苄酯", "鄰苯二甲酸丁苄酯", "鄰苯二甲酸苯基丁酯"]),
    ("DBP", ["dbp", "dibutyl phthalate", "di-butyl phthalate", "di-n-butyl phthalate", "邻苯二甲酸二丁酯", "鄰苯二甲酸二丁酯"]),
    ("DEHP", ["dehp", "bis(2-ethylhexyl) phthalate", "di(2-ethylhexyl) phthalate", "di-2-ethyl hexyl phthalate", "di-(2-ethylhexyl) phthalate", "邻苯二甲酸二(2-乙基", "鄰苯二甲酸二(2-乙基", "邻苯二甲酸二异辛酯", "鄰苯二甲酸二異辛酯"]),
    ("DIBP", ["dibp", "diisobutyl phthalate", "diisobutyl phthalates", "di-(iso-butyl) phthalate", "邻苯二甲酸二异丁酯", "鄰苯二甲酸二異丁酯"]),
]

ROHS_DIRECTIVE_KEYWORDS = [
    "rohs",
    "2011/65/eu",
    "2015/863",
    "lead",
    "cadmium",
    "mercury",
    "hexavalent chromium",
    "pbb",
    "pbde",
    "dehp",
    "bbp",
    "dbp",
    "dibp",
    "铅",
    "鉛",
    "镉",
    "鎘",
    "汞",
    "六价铬",
    "六價鉻",
    "多溴联苯",
    "多溴聯苯",
    "多溴二苯醚",
    "多溴聯苯醚",
    "邻苯",
    "鄰苯",
]

RESULT_TOKEN = re.compile(
    r"(?i)(?:n\.?\s*d\.?|not\s+detected|pass|negative|<\s*\d+(?:\.\d+)?\s*(?:mg/kg|ppm|%|ug/g|µg/g)?|\d+(?:\.\d+)?\s*(?:mg/kg|ppm|%|ug/g|µg/g))"
)
ROW_VALUE_TOKEN = re.compile(r"(?i)n\.?\s*d\.?|not\s+detected|pass|negative|<\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?")


@dataclass
class PageText:
    page: int
    method: str
    char_count: int
    text: str = ""
    error: str = ""


@dataclass
class PdfRecord:
    file_name: str
    file_path: str
    file_size: int
    page_count: int = 0
    encrypted: bool | None = None
    text_char_count: int = 0
    extraction_status: str = "not_started"
    needs_ocr: bool = False
    needs_pdf_fallback: bool = False
    lab_name: str = ""
    report_no: str = ""
    report_date: str = ""
    sample_name: str = ""
    test_part: str = ""
    substances: dict[str, str] = field(default_factory=dict)
    confidence: float = 0.0
    is_rohs_related: bool | None = None
    rohs_confidence: float = 0.0
    non_rohs_confidence: float = 0.0
    result_confidence: float = 0.0
    ai_assisted: bool = False
    ai_model: str = ""
    ai_fields: list[str] = field(default_factory=list)
    segment_id: str = ""
    notes: list[str] = field(default_factory=list)
    pages: list[PageText] = field(default_factory=list)


@dataclass
class AiConfig:
    enabled: bool = True
    provider: str = "gemini"
    api_key: str = ""
    nvidia_api_key: str = ""
    model: str = "gemini-3.1-flash-lite"
    low_confidence_threshold: float = 0.6
    timeout_seconds: int = 45
    log_path: Path | None = DEFAULT_AI_LOG
    log_files_dir: Path | None = DEFAULT_AI_LOG_FILES


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def safe_sheet_value(value):
    if isinstance(value, list):
        return "; ".join(str(x) for x in value)
    return value


def load_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key:
            continue
        if (
            len(value) >= 2
            and value[0] == value[-1]
            and value[0] in {"'", '"'}
        ):
            value = value[1:-1]
        values[key] = value
    return values


def env_value(name: str, env_file_values: dict[str, str]) -> str:
    return env_file_values.get(name) or os.environ.get(name) or ""


def truncate_text(value: str, max_chars: int) -> str:
    text = value or ""
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n...[truncated]"


def parse_json_object(value: str) -> dict:
    text = normalize_spaces(value)
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.I)
    text = re.sub(r"\s*```$", "", text)
    try:
        parsed = json.loads(text)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        pass
    match = re.search(r"\{.*\}", text, flags=re.S)
    if not match:
        return {}
    try:
        parsed = json.loads(match.group(0))
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        return {}


def call_gemini_json(prompt: str, config: AiConfig) -> dict:
    endpoint = (
        "https://generativelanguage.googleapis.com/v1beta/"
        f"models/{config.model}:generateContent"
    )
    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [{"text": prompt}],
            }
        ],
        "generationConfig": {
            "temperature": 0,
            "responseMimeType": "application/json",
        },
    }
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        endpoint,
        data=body,
        headers={
            "Content-Type": "application/json",
            "x-goog-api-key": config.api_key,
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=config.timeout_seconds) as response:
        data = json.loads(response.read().decode("utf-8"))
    parts = (
        data.get("candidates", [{}])[0]
        .get("content", {})
        .get("parts", [])
    )
    text = "\n".join(part.get("text", "") for part in parts if isinstance(part, dict))
    return parse_json_object(text)


def call_nvidia_json(prompt: str, config: AiConfig) -> dict:
    endpoint = "https://integrate.api.nvidia.com/v1/chat/completions"
    payload = {
        "model": config.model,
        "messages": [
            {
                "role": "user",
                "content": prompt,
            }
        ],
        "temperature": 0,
        "max_tokens": 4096,
        "stream": False,
    }
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        endpoint,
        data=body,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {config.nvidia_api_key}",
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=config.timeout_seconds) as response:
        data = json.loads(response.read().decode("utf-8"))
    text = (
        data.get("choices", [{}])[0]
        .get("message", {})
        .get("content", "")
    ) or ""
    return parse_json_object(text)


def call_ai_json(prompt: str, config: AiConfig) -> dict:
    if config.provider == "nvidia":
        return call_nvidia_json(prompt, config)
    return call_gemini_json(prompt, config)


def ai_log_snapshot(record: PdfRecord) -> dict:
    substances = {}
    for name, _aliases in SUBSTANCES:
        value, unit = result_cells(record, name)
        substances[name] = {
            "value": value,
            "unit": unit,
            "source": substance_source(record, name),
            "confidence": round(substance_confidence(record, name), 2),
            "evidence": truncate_text(record.substances.get(f"{name} Evidence", ""), 500),
        }
    return {
        "report_no": record.report_no,
        "report_date": record.report_date,
        "confidence": record.confidence,
        "result_confidence": record.result_confidence,
        "ai_fields": list(record.ai_fields),
        "substances": substances,
    }


def diff_ai_snapshots(before: dict, after: dict) -> dict:
    changes = {}
    for key in ["report_no", "report_date", "confidence", "result_confidence", "ai_fields"]:
        if before.get(key) != after.get(key):
            changes[key] = {"before": before.get(key), "after": after.get(key)}
    substance_changes = {}
    before_substances = before.get("substances", {})
    after_substances = after.get("substances", {})
    for name in sorted(set(before_substances) | set(after_substances)):
        if before_substances.get(name) != after_substances.get(name):
            substance_changes[name] = {
                "before": before_substances.get(name),
                "after": after_substances.get(name),
            }
    if substance_changes:
        changes["substances"] = substance_changes
    return changes


def write_ai_log(
    record: PdfRecord,
    config: AiConfig,
    purpose: str,
    trigger_reason: str,
    targeted_fields: list[str],
    snippets: str,
    ai_response: dict | None,
    before: dict,
    after: dict,
    error: str = "",
    copied_source_file: str = "",
) -> None:
    if not config.log_path:
        return
    config.log_path.parent.mkdir(parents=True, exist_ok=True)
    entry = {
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "file_name": record.file_name,
        "file_path": record.file_path,
        "segment_id": record.segment_id,
        "report_no_before": before.get("report_no"),
        "report_no_after": after.get("report_no"),
        "purpose": purpose,
        "trigger_reason": trigger_reason,
        "targeted_fields": targeted_fields,
        "provider": config.provider,
        "model": config.model,
        "rohs_confidence": record.rohs_confidence,
        "non_rohs_confidence": record.non_rohs_confidence,
        "result_confidence_before": before.get("result_confidence"),
        "result_confidence_after": after.get("result_confidence"),
        "changes": diff_ai_snapshots(before, after),
        "ai_response": ai_response or {},
        "error": error,
        "copied_source_file": copied_source_file,
        "snippet_preview": truncate_text(snippets, 3000),
    }
    with config.log_path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(entry, ensure_ascii=False) + "\n")


def safe_filename_part(value: str, fallback: str = "unknown") -> str:
    safe = re.sub(r"[^A-Za-z0-9_.()-]+", "_", value or "").strip("._")
    return safe or fallback


def copy_ai_log_source_file(record: PdfRecord, config: AiConfig, purpose: str) -> str:
    if purpose not in {"low_confidence", "fallback_substances"}:
        return ""
    if record.is_rohs_related is not True:
        return ""
    if not config.log_files_dir:
        return ""
    source = Path(record.file_path)
    if not source.exists():
        return ""
    config.log_files_dir.mkdir(parents=True, exist_ok=True)
    stem = safe_filename_part(source.stem)
    report_no = safe_filename_part(record.report_no, "no_report_no")
    segment = safe_filename_part(record.segment_id.replace("/", "_") if record.segment_id else "single")
    dest = config.log_files_dir / f"{stem}__{report_no}__segment_{segment}{source.suffix}"
    if not dest.exists():
        shutil.copy2(source, dest)
    return str(dest)


def interesting_lines(text: str, keywords: list[str], context: int = 2, max_lines: int = 70) -> list[str]:
    lines = [normalize_spaces(line) for line in text.splitlines()]
    lines = [line for line in lines if line]
    if not lines:
        return []
    indexes: set[int] = set()
    low_keywords = [keyword.lower() for keyword in keywords]
    for idx, line in enumerate(lines):
        low_line = line.lower()
        if any(keyword in low_line for keyword in low_keywords):
            for nearby in range(max(0, idx - context), min(len(lines), idx + context + 1)):
                indexes.add(nearby)
    if not indexes:
        return lines[:max_lines]
    selected = [lines[idx] for idx in sorted(indexes)]
    return selected[:max_lines]


def collect_ai_snippets(record: PdfRecord, purpose: str, max_chars: int = 14000) -> str:
    identity_keywords = [
        "report no",
        "test report no",
        "number",
        "date",
        "issue date",
        "報告",
        "編號",
        "日期",
    ]
    result_keywords = [
        "rohs",
        "2011/65/eu",
        "2015/863",
        "test result",
        "test item",
        "restricted substances",
        "lead",
        "cadmium",
        "mercury",
        "hexavalent",
        "chromium",
        "pbb",
        "pbde",
        "dehp",
        "bbp",
        "dbp",
        "dibp",
        "limit",
        "mdl",
        "unit",
        "pass",
        "nd",
        "n.d.",
    ]
    keywords = identity_keywords if purpose == "identity" else identity_keywords + result_keywords
    chunks = []
    if purpose == "identity":
        for page in record.pages[:2]:
            if not page.text:
                continue
            lines = [normalize_spaces(line) for line in page.text.splitlines() if normalize_spaces(line)]
            if lines:
                head_tail = lines[:35]
                if len(lines) > 35:
                    head_tail += ["..."] + lines[-35:]
                chunks.append(
                    f"[Page {page.page}, method={page.method}, first and last lines]\n"
                    + "\n".join(head_tail)
                )

    scored_pages = []
    for page in record.pages:
        if not page.text:
            continue
        text = page.text
        low = text.lower()
        score = 0
        if page.page <= 3:
            score += 4
        for keyword in keywords:
            if keyword.lower() in low:
                score += 1
        if purpose == "low_confidence" and re.search(r"(?i)test\s+item|test\s+result|mdl|limit|n\.?\s*d\.?", text):
            score += 5
        if score:
            scored_pages.append((score, page.page, page))
    if not scored_pages:
        scored_pages = [(1, page.page, page) for page in record.pages[:3] if page.text]
    scored_pages.sort(key=lambda item: (-item[0], item[1]))

    used_pages: set[int] = set()
    for _score, _page_no, page in scored_pages:
        if page.page in used_pages:
            continue
        used_pages.add(page.page)
        lines = interesting_lines(page.text, keywords, context=2, max_lines=80)
        if not lines:
            continue
        chunk = (
            f"[Page {page.page}, method={page.method}]\n"
            + "\n".join(lines)
        )
        chunks.append(chunk)
        combined = "\n\n".join(chunks)
        if len(combined) >= max_chars:
            return truncate_text(combined, max_chars)
        if purpose == "identity" and len(used_pages) >= 4:
            break
        if purpose == "low_confidence" and len(used_pages) >= 6:
            break
    return truncate_text("\n\n".join(chunks), max_chars)


def ai_record_snapshot(record: PdfRecord) -> dict:
    substances = {}
    for name, _aliases in SUBSTANCES:
        value, unit = result_cells(record, name)
        substances[name] = {
            "value": value,
            "unit": unit,
            "source": substance_source(record, name),
            "substance_confidence": round(substance_confidence(record, name), 2),
            "evidence": record.substances.get(f"{name} Evidence", ""),
        }
    return {
        "file_name": record.file_name,
        "report_no": record.report_no,
        "report_date": record.report_date,
        "report_date_note": "If blank, use the priority in the task instructions instead of filename.",
        "lab_name": record.lab_name,
        "sample_name": record.sample_name,
        "test_part": record.test_part,
        "rohs_confidence": record.rohs_confidence,
        "non_rohs_confidence": record.non_rohs_confidence,
        "result_confidence": record.result_confidence,
        "status": record.extraction_status,
        "substances": substances,
        "notes": record.notes,
    }


def build_ai_prompt(record: PdfRecord, purpose: str) -> str:
    snippets = collect_ai_snippets(record, purpose)
    snapshot = json.dumps(ai_record_snapshot(record), ensure_ascii=False, indent=2)
    schema = {
        "report_no": "string or null",
        "report_date": "YYYY-MM-DD string or null",
        "date_source": "one of: report_date, issue_date, approval_date, testing_period_end, sample_testing_date_end, sample_received_date, null",
        "substances": {
            name: {"value": "string or null", "unit": "string or null", "evidence": "short text or null"}
            for name, _aliases in SUBSTANCES
        },
        "confidence": "number 0-1",
        "evidence": "short text showing the matched phrase",
        "reason": "short explanation",
    }
    if purpose == "identity":
        task = (
            "Find only the report number and date. Do not infer from filename. "
            "Date priority: explicit report date/issue date/approval date/signature date first. "
            "For CTI reports, Chinese '日 期' near the signature/seal area is the report date. If absent, use the "
            "testing period end date or sample testing date end date. If that is absent, use sample "
            "received date. Set date_source to the source used."
        )
    elif purpose == "fallback_substances":
        fuzzy = ", ".join(fallback_ambiguous_substances(record))
        task = (
            "Verify only these fallback-derived substances: "
            f"{fuzzy}. Do not change other substances. Use the snippets to find whether each target has "
            "a direct result table value. Treat MDL, Limit, RoHS limit, method detection limit, and "
            "specification limits as non-results. If a direct result is visible, return it. If no direct "
            "result is visible in the snippets, return null for that substance."
        )
    else:
        task = (
            "Compare the extraction snapshot against the report snippets. Correct only fields "
            "that are visible in the snippets. Treat MDL, Limit, RoHS limit, method detection "
            "limit, and specification limits as non-results. Return ND only when the sample "
            "result is explicitly ND/N.D./Not Detected. If report_date is blank, also fill it "
            "using this priority: explicit report/issue/approval/signature date, testing period end date, "
            "sample testing date end date, sample received date."
        )
    return (
        "You are assisting RoHS PDF extraction. Use ONLY the provided targeted snippets; "
        "do not assume you have the full document. If a value is not visible, return null.\n"
        f"Task: {task}\n\n"
        "Return JSON only with this schema:\n"
        f"{json.dumps(schema, ensure_ascii=False, indent=2)}\n\n"
        "Current extraction snapshot:\n"
        f"{snapshot}\n\n"
        "Targeted report snippets:\n"
        f"{snippets}"
    )


def should_ai_check_identity(record: PdfRecord) -> bool:
    if record.is_rohs_related is not True:
        return False
    note_text = " ".join(record.notes)
    return (
        not record.report_no
        or not record.report_date
        or "報告號碼未能從PDF內容判讀" in note_text
        or "報告日期未能從PDF內容判讀" in note_text
    )


def should_ai_check_low_confidence(record: PdfRecord, config: AiConfig) -> bool:
    return record.is_rohs_related is True and record.confidence < config.low_confidence_threshold


def substance_source(record: PdfRecord, name: str) -> str:
    evidence = record.substances.get(f"{name} Evidence", "")
    low = evidence.lower()
    if not record.substances.get(name):
        return "missing"
    if evidence.startswith("AI判定"):
        return "ai"
    if evidence.startswith("SGS Taiwan split result table"):
        return "structured_table"
    if "fallback: sgs taiwan rohs summary" in low:
        return "summary_fallback"
    if "fallback: substance row contains" in low:
        return "nearby_nd_fallback"
    if "fallback: rohs phthalate" in low:
        return "summary_fallback"
    if evidence:
        return "rule_table"
    return "unknown"


def substance_confidence(record: PdfRecord, name: str) -> float:
    source = substance_source(record, name)
    weights = {
        "missing": 0.0,
        "ai": 0.9,
        "structured_table": 0.95,
        "rule_table": 0.85,
        "summary_fallback": 0.55,
        "nearby_nd_fallback": 0.45,
        "unknown": 0.55,
    }
    return weights.get(source, 0.5)


def text_for_record(record: PdfRecord) -> str:
    return "\n".join(page.text for page in record.pages if page.text)


def has_result_table_clues_for_substance(record: PdfRecord, name: str) -> bool:
    text = text_for_record(record)
    if not text:
        return False
    compact = normalize_compact_text(text)
    low = compact.lower()
    aliases = dict(SUBSTANCES).get(name, [])
    has_alias = any(alias.lower() in low or squash_for_keyword(alias) in squash_for_keyword(compact) for alias in aliases)
    if not has_alias:
        return False
    has_result_context = bool(
        re.search(
            r"(?i)(test\s*item|test\s*result|測試項目|測試結果|检测项目|检测结果|"
            r"\bmdl\b|method\s+detection|限值|limit|mg/kg|ppm)",
            compact,
        )
    )
    has_result_token = bool(re.search(r"(?i)n\.?\s*d\.?|not\s+detected|<\s*\d|\d+(?:\.\d+)?\s*(?:mg/kg|ppm|%)", compact))
    return has_result_context and has_result_token


def fallback_ambiguous_substances(record: PdfRecord) -> list[str]:
    fuzzy = []
    for name, _aliases in SUBSTANCES:
        source = substance_source(record, name)
        if source == "nearby_nd_fallback" and has_result_table_clues_for_substance(record, name):
            fuzzy.append(name)
    return fuzzy


def should_ai_check_fallback_substances(record: PdfRecord) -> bool:
    return record.is_rohs_related is True and bool(fallback_ambiguous_substances(record))


def mark_ai_field(record: PdfRecord, field_name: str) -> None:
    record.ai_assisted = True
    if field_name not in record.ai_fields:
        record.ai_fields.append(field_name)


def apply_ai_response(record: PdfRecord, payload: dict, purpose: str) -> None:
    if not payload:
        return
    note_text = " ".join(record.notes)
    report_no_needs_ai = not record.report_no or "報告號碼未能從PDF內容判讀" in note_text
    report_date_needs_ai = not record.report_date or "報告日期未能從PDF內容判讀" in note_text
    report_no = normalize_spaces(str(payload.get("report_no") or ""))
    if report_no and (purpose == "low_confidence" or report_no_needs_ai):
        cleaned = clean_report_no(report_no)
        if is_plausible_report_no(cleaned):
            record.report_no = cleaned
            mark_ai_field(record, "報告號碼")

    report_date = normalize_spaces(str(payload.get("report_date") or ""))
    if report_date and (purpose == "low_confidence" or report_date_needs_ai):
        normalized = normalize_report_date(report_date)
        if normalized:
            record.report_date = normalized
            mark_ai_field(record, "報告日期")
            date_source = normalize_spaces(str(payload.get("date_source") or ""))
            if date_source:
                record.notes.append(f"AI日期來源：{date_source}。")

    substances = payload.get("substances")
    if purpose in {"low_confidence", "fallback_substances"} and isinstance(substances, dict):
        allowed_names = (
            set(fallback_ambiguous_substances(record))
            if purpose == "fallback_substances"
            else {name for name, _aliases in SUBSTANCES}
        )
        for name, _aliases in SUBSTANCES:
            if name not in allowed_names:
                continue
            item = substances.get(name)
            if not isinstance(item, dict):
                continue
            value = normalize_spaces(str(item.get("value") or ""))
            unit = normalize_spaces(str(item.get("unit") or ""))
            if not value:
                continue
            combined = normalize_spaces(f"{value} {unit}") if unit and unit.lower() not in value.lower() else value
            record.substances[name] = combined
            evidence = normalize_spaces(str(item.get("evidence") or payload.get("evidence") or ""))
            record.substances[f"{name} Evidence"] = f"AI判定: {evidence}" if evidence else "AI判定"
            mark_ai_field(record, name)

    if record.ai_assisted:
        record.notes.append(
            f"AI判定：{', '.join(record.ai_fields)}；模型={record.ai_model or 'gemini-3.1-flash-lite'}。"
        )


def refresh_record_confidence(record: PdfRecord) -> None:
    filled_core = sum(bool(x) for x in [record.report_no, record.report_date, record.lab_name, record.sample_name])
    substance_score = sum(substance_confidence(record, name) for name, _aliases in SUBSTANCES) / len(SUBSTANCES)
    record.result_confidence = round((filled_core / 4) * 0.35 + substance_score * 0.65, 2)
    if record.is_rohs_related is True:
        record.confidence = record.result_confidence
    elif record.is_rohs_related is False:
        record.confidence = record.non_rohs_confidence
    else:
        record.confidence = max(record.rohs_confidence, record.non_rohs_confidence)


def apply_ai_assistance(record: PdfRecord, config: AiConfig) -> None:
    identity_needed = should_ai_check_identity(record)
    low_confidence_needed = should_ai_check_low_confidence(record, config)
    fallback_substances_needed = should_ai_check_fallback_substances(record)
    if not (identity_needed or low_confidence_needed or fallback_substances_needed):
        return
    if not config.enabled:
        return
    if config.provider == "nvidia" and not config.nvidia_api_key:
        record.notes.append("AI輔助未執行：NVIDIA provider 缺少 NVIDIA_API_KEY。")
        return
    if config.provider != "nvidia" and not config.api_key:
        record.notes.append("AI輔助未執行：目前 Gemini provider 缺少 GEMINI_API_KEY 或 GOOGLE_API_KEY。")
        return

    purposes = []
    if identity_needed:
        purposes.append("identity")
    if low_confidence_needed:
        purposes.append("low_confidence")
    elif fallback_substances_needed:
        purposes.append("fallback_substances")

    for purpose in purposes:
        fuzzy_substances = fallback_ambiguous_substances(record)
        if purpose == "identity":
            targeted_fields = ["報告號碼", "報告日期"]
            trigger_reason = "RoHS report has missing/fallback report number or report date."
        elif purpose == "low_confidence":
            targeted_fields = [name for name, _aliases in SUBSTANCES]
            trigger_reason = f"RoHS report confidence {record.confidence:.2f} is below threshold {config.low_confidence_threshold:.2f}."
        else:
            targeted_fields = fuzzy_substances
            trigger_reason = "Fallback-derived nearby ND result is ambiguous and result-table clues exist."
        snippets = collect_ai_snippets(
            record,
            "low_confidence" if purpose in {"low_confidence", "fallback_substances"} else "identity",
        )
        copied_source_file = copy_ai_log_source_file(record, config, purpose)
        before_snapshot = ai_log_snapshot(record)
        try:
            payload = call_ai_json(build_ai_prompt(record, purpose), config)
        except (urllib.error.URLError, TimeoutError, OSError, json.JSONDecodeError) as exc:
            record.notes.append(f"AI輔助失敗：{type(exc).__name__}: {exc}")
            write_ai_log(
                record=record,
                config=config,
                purpose=purpose,
                trigger_reason=trigger_reason,
                targeted_fields=targeted_fields,
                snippets=snippets,
                ai_response=None,
                before=before_snapshot,
                after=ai_log_snapshot(record),
                error=f"{type(exc).__name__}: {exc}",
                copied_source_file=copied_source_file,
            )
            if copied_source_file:
                record.notes.append(f"AI輔助低信心原檔已保存：{copied_source_file}")
            continue
        record.ai_model = config.model
        before_fields = set(record.ai_fields)
        apply_ai_response(record, payload, purpose)
        if set(record.ai_fields) != before_fields:
            refresh_record_confidence(record)
        write_ai_log(
            record=record,
            config=config,
            purpose=purpose,
            trigger_reason=trigger_reason,
            targeted_fields=targeted_fields,
            snippets=snippets,
            ai_response=payload,
            before=before_snapshot,
            after=ai_log_snapshot(record),
            copied_source_file=copied_source_file,
        )
        if copied_source_file:
            record.notes.append(f"AI輔助低信心原檔已保存：{copied_source_file}")


def extract_text_with_pypdf(pdf_path: Path) -> PdfRecord:
    record = PdfRecord(
        file_name=pdf_path.name,
        file_path=str(pdf_path),
        file_size=pdf_path.stat().st_size,
    )

    try:
        reader = PdfReader(str(pdf_path))
        record.encrypted = bool(getattr(reader, "is_encrypted", False))
        record.page_count = len(reader.pages)
    except Exception as exc:
        record.extraction_status = "reader_failed"
        record.needs_pdf_fallback = True
        record.notes.append(f"PdfReader failed: {type(exc).__name__}: {exc}")
        return record

    for idx in range(record.page_count):
        try:
            text = reader.pages[idx].extract_text() or ""
            record.pages.append(
                PageText(page=idx + 1, method="pypdf", char_count=len(text), text=text)
            )
        except Exception as exc:
            record.pages.append(
                PageText(
                    page=idx + 1,
                    method="pypdf",
                    char_count=0,
                    error=f"{type(exc).__name__}: {exc}",
                )
            )
            record.notes.append(f"Page {idx + 1}: {type(exc).__name__}: {exc}")

    record.text_char_count = sum(page.char_count for page in record.pages)
    failed_pages = sum(1 for page in record.pages if page.error)

    if record.text_char_count >= 500:
        record.extraction_status = "text_extracted"
    elif failed_pages:
        record.extraction_status = "text_failed_or_encrypted"
        record.needs_pdf_fallback = True
        record.needs_ocr = True
    else:
        record.extraction_status = "low_or_no_text"
        record.needs_ocr = True

    combined_text = "\n".join(page.text for page in record.pages)
    if combined_text and is_probably_garbled_text(combined_text):
        ocr_pages = extract_text_with_ocr(pdf_path, record.page_count)
        if ocr_pages:
            record.pages.extend(ocr_pages)
            record.text_char_count = sum(page.char_count for page in record.pages)
            record.extraction_status = "text_extracted"
            record.notes.append("Garbled PDF text layer detected; OCR text was added.")

    return record


def is_probably_garbled_text(text: str) -> bool:
    if not text:
        return False
    cid_hits = len(re.findall(r"\(cid:\d+\)|/i\d+", text))
    replacement_hits = text.count("\ufffd")
    ascii_letters = len(re.findall(r"[A-Za-z]", text))
    total = max(len(text), 1)
    return cid_hits >= 20 or replacement_hits >= 20 or (ascii_letters / total < 0.08 and cid_hits >= 5)


def extract_text_with_ocr(pdf_path: Path, page_count: int, max_pages: int = 8) -> list[PageText]:
    try:
        import fitz
        from PIL import Image
        import pytesseract
    except Exception:
        return []

    pages: list[PageText] = []
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return []

    for idx in range(min(page_count, max_pages)):
        try:
            page = doc[idx]
            pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
            image = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            text = pytesseract.image_to_string(image, lang="eng") or ""
            pages.append(
                PageText(page=idx + 1, method="ocr", char_count=len(text), text=text)
            )
        except Exception as exc:
            pages.append(
                PageText(
                    page=idx + 1,
                    method="ocr",
                    char_count=0,
                    error=f"{type(exc).__name__}: {exc}",
                )
            )
    return pages


def extract_text_with_markitdown(pdf_path: Path) -> str:
    try:
        from markitdown import MarkItDown
    except Exception:
        return ""
    try:
        result = MarkItDown(enable_plugins=False).convert(str(pdf_path))
    except Exception:
        return ""
    return getattr(result, "text_content", "") or getattr(result, "markdown", "") or ""


def record_has_ocr_text(record: PdfRecord) -> bool:
    return any(page.method == "ocr" for page in record.pages)


def record_has_markitdown_text(record: PdfRecord) -> bool:
    return any(page.method == "markitdown" for page in record.pages)


def add_markitdown_text(record: PdfRecord, reason: str) -> bool:
    if record_has_markitdown_text(record):
        return False
    text = extract_text_with_markitdown(Path(record.file_path))
    if not text:
        record.notes.append(f"MarkItDown fallback attempted but no usable text was extracted. Reason: {reason}")
        return False
    page_no = max([page.page for page in record.pages] or [0]) + 1
    record.pages.append(PageText(page=page_no, method="markitdown", char_count=len(text), text=text))
    record.text_char_count = sum(page.char_count for page in record.pages)
    record.extraction_status = "text_extracted"
    record.notes.append(f"MarkItDown text was added. Reason: {reason}")
    return True


def add_ocr_text(record: PdfRecord, reason: str, max_pages: int | None = None) -> bool:
    if record_has_ocr_text(record):
        return False
    ocr_pages = extract_text_with_ocr(
        Path(record.file_path),
        record.page_count,
        max_pages=max_pages or 8,
    )
    usable_pages = [page for page in ocr_pages if page.text]
    if not usable_pages:
        record.notes.append(f"OCR fallback attempted but no usable text was extracted. Reason: {reason}")
        return False
    record.pages.extend(ocr_pages)
    record.text_char_count = sum(page.char_count for page in record.pages)
    record.extraction_status = "text_extracted"
    record.needs_ocr = False
    record.notes.append(f"OCR text was added. Reason: {reason}")
    return True


def detect_lab(text: str) -> str:
    low = text.lower()

    lab_patterns = [
        ("SGS", [r"\bsgs\b", r"sgs-cstc", r"台灣檢驗科技", r"通标标准", r"通標標準"]),
        ("Intertek", [r"\bintertek\b", r"全國公證"]),
        ("BV", [r"\bbureau\s+veritas\b", r"\bbv\b", r"必维", r"必維"]),
        ("TUV", [r"\btuv\b", r"\btüv\b", r"tuv\s+rheinland", r"tüv\s+rheinland"]),
        ("PONY", [r"\bpony\b", r"谱尼", r"譜尼", r"青岛谱尼", r"青島譜尼"]),
        ("SAFETY", [r"guangdong\s+safety\s+testing", r"\bsafety\s+testing\b", r"广东斯富特", r"廣東斯富特", r"\bsft\b"]),
        ("CTI", [r"centre\s+testing\s+international", r"\bcti\b", r"华测检测", r"華測檢測", r"华测检测认证", r"華測檢測認證"]),
        ("NTEK", [r"\bntek\b", r"北测检测", r"北測檢測"]),
    ]

    for name, patterns in lab_patterns:
        if any(re.search(pattern, low, re.IGNORECASE) for pattern in patterns):
            return name
    return ""


def assess_rohs_relevance(text: str) -> tuple[bool | None, float, float]:
    relevant = rohs_relevant_text(text)
    compact = normalize_spaces(relevant)
    if not compact:
        return None, 0.0, 0.0

    low = relevant.lower()
    squashed = squash_for_keyword(relevant)

    directive_patterns = [
        r"rohs\s+directive",
        r"directive\s*\(eu\)\s*2015/863",
        r"2011/65/eu",
        r"2015/863",
        r"restriction\s+of\s+the\s+use\s+of\s+certain\s+hazardous",
        r"iec\s*62321",
        r"en\s*62321",
    ]
    directive_hits = sum(1 for pattern in directive_patterns if re.search(pattern, relevant, re.IGNORECASE))

    requested_pattern = (
        r"(?i)(test\s*(?:requested|requirement|item|result)|determine|to\s+test)"
        r".{0,220}?"
        r"(lead|cadmium|mercury|hexavalent|chromium\s*vi|pbb|pbde|phthalate|dehp|bbp|dbp|dibp)"
    )
    has_requested_context = bool(re.search(requested_pattern, relevant))

    row_hit_count = len(rohs_result_row_hits(relevant))
    hits = sum(1 for keyword in ROHS_DIRECTIVE_KEYWORDS if keyword.lower() in low)
    squashed_hits = sum(1 for keyword in ROHS_DIRECTIVE_KEYWORDS if squash_for_keyword(keyword) in squashed)
    keyword_hits = max(hits, squashed_hits)
    substance_hits = 0
    for _name, aliases in SUBSTANCES:
        if any(alias.lower() in low or squash_for_keyword(alias) in squashed for alias in aliases):
            substance_hits += 1

    rohs_score = 0.0
    if re.search(r"(?i)\brohs\b", relevant):
        rohs_score += 0.2
    rohs_score += min(directive_hits, 3) * 0.12
    if has_requested_context:
        rohs_score += 0.14
    rohs_score += min(row_hit_count / len(SUBSTANCES), 1.0) * 0.32
    rohs_score += min(substance_hits / len(SUBSTANCES), 1.0) * 0.22
    if keyword_hits >= 3:
        rohs_score += 0.12
    rohs_score = min(rohs_score, 1.0)

    alternate_test_patterns = [
        r"(?i)red\s+phosphorus",
        r"(?i)\bhalogen\b",
        r"(?i)\breach\b",
        r"(?i)\bsvhc\b",
        r"(?i)phosphorus",
        r"(?i)flame\s+retardant",
    ]
    has_alternate_test = any(re.search(pattern, relevant) for pattern in alternate_test_patterns)
    has_report_context = bool(
        re.search(r"(?i)(test\s+report|report\s+no|sample\s+name|date)", relevant)
    )

    non_rohs_score = 0.0
    if not re.search(r"(?i)\brohs\b|2011/65/eu|2015/863|iec\s*62321|en\s*62321", relevant):
        non_rohs_score += 0.35
    if row_hit_count == 0:
        non_rohs_score += 0.25
    if substance_hits < 2:
        non_rohs_score += 0.15
    if has_alternate_test:
        non_rohs_score += 0.25
    if has_report_context and len(compact) >= 300:
        non_rohs_score += 0.1
    non_rohs_score = min(non_rohs_score, 1.0)

    if rohs_score >= 0.55 and rohs_score >= non_rohs_score:
        decision: bool | None = True
    elif non_rohs_score >= 0.75 and rohs_score < 0.45:
        decision = False
    else:
        decision = None
    return decision, round(rohs_score, 2), round(non_rohs_score, 2)


def is_rohs_report(text: str) -> bool:
    decision, _rohs_confidence, _non_rohs_confidence = assess_rohs_relevance(text)
    return decision is True


def rohs_relevant_text(text: str) -> str:
    excluded_line = re.compile(
        r"(?i)(sample\s*(?:name|description)|product\s*name|model|style|p\.?\s*o\.?\s*no|buyer|manufacturer)"
    )
    lines = []
    for line in text.splitlines():
        clean = normalize_spaces(line)
        if not clean:
            continue
        if excluded_line.search(clean):
            continue
        lines.append(clean)
    return "\n".join(lines)


def rohs_result_row_hits(text: str) -> set[str]:
    hits: set[str] = set()
    row_pattern = re.compile(
        r"(?i)(lead\s*\(pb\)|cadmium\s*\(cd\)|mercury\s*\(hg\)|"
        r"hexavalent\s+chromium|chromium\s*vi|cr\s*\(?vi\)?|"
        r"pbb(?:s)?\b|pbde(?:s)?\b|dehp|bbp|dbp|dibp|phthalate)"
        r".{0,140}?"
        r"(?:n\.?\s*d\.?|not\s+detected|pass|<\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?\s*(?:mg/kg|ppm|%|ug/cm2|ug/g))"
    )
    for match in row_pattern.finditer(text):
        hits.add(match.group(1).lower())
    return hits


def squash_for_keyword(text: str) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", text.lower())


def first_match(patterns: Iterable[str], text: str) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if match:
            return normalize_spaces(match.group(1))
    return ""


def first_line_match(patterns: Iterable[str], text: str) -> str:
    for line in text.splitlines():
        clean = normalize_spaces(line)
        for pattern in patterns:
            match = re.search(pattern, clean, re.IGNORECASE)
            if match:
                return normalize_spaces(match.group(1))
    return ""


def first_pattern_match(patterns: Iterable[str], text: str) -> str:
    lines = [normalize_spaces(line) for line in text.splitlines()]
    for pattern in patterns:
        for clean in lines:
            match = re.search(pattern, clean, re.IGNORECASE)
            if match:
                return normalize_spaces(match.group(1))
    return ""


def extract_report_no(text: str, file_stem: str) -> str:
    lines = [normalize_spaces(line) for line in text.splitlines() if normalize_spaces(line)]
    direct_patterns = [
        r"(?:Test\s+Report\s+No\.?|Report\s*No\.?)\s*[:：]?\s*([A-Z0-9][A-Z0-9_\-()/\.]{4,})",
        r"\bLAB(?:\s*NO\.?)?\s*(?:[:：]|\d+)?\s*(\([0-9]{3,}\)[A-Z0-9_\-()/\.]{4,}|[A-Z0-9(][A-Z0-9_\-()/\.]{4,})",
        r"(?:报告编号|報告編號)\s*(?:Report\s*No\.?)?\s*[:：]?\s*([A-Z0-9][A-Z0-9_\-()/\.]{4,})",
        r"(?:Report\s+Number|报告号|報告號)\s*\(?\s*REPORT\s*NO\.?\s*\)?\s*[:：]?\s*([A-Z0-9(][A-Z0-9_\-()/\.]{4,})",
        r"(?:報告\s*號碼|报告\s*号码|報告\s*编号|报告\s*編號)\s*[:：]\s*([A-Z0-9(][A-Z0-9_\-()/\.]{4,})",
        r"(?:報告\s*號碼|报告\s*号码|號碼|号码)\s*\(?\s*No\.?\s*\)?\s*[:：]?\s*([A-Z0-9(][A-Z0-9_\-()/\.]{4,})",
        r"(?:测试报告|測試報告|試驗報告|检测报告|檢測報告)\s*\.?\s*No\.?\s*[:：]?\s*([A-Z0-9][A-Z0-9_\-()/\.]{4,})",
        r"^No\.?\s*[:：]\s*([A-Z0-9][A-Z0-9_\-()/\.]{4,})(?:\s+Report\b|\s*$)",
        r"\bNo\.?\s*[:：]\s*([A-Z0-9][A-Z0-9_\-()/\.]{4,})(?:\s+Date\b|\s+日期|\s*Test\s+Report\b)",
        r"\bNo\.?\s+([A-Z]{2,}[A-Z0-9_\-()/\.]{4,})\s+Date\b",
    ]

    for line in lines[:900]:
        if is_historical_report_no_context(line):
            continue
        for candidate_line in [line, repair_spaced_report_line(line)]:
            loose = re.search(r"(?i)(?:Test\s+Report\s+No\.?|Report\s*No\.?)\s*[:：]\s*(.+?)(?:\s*Date\b|\s*日期|$)", candidate_line)
            if loose:
                candidate = clean_report_no(loose.group(1))
                if is_plausible_report_no(candidate):
                    return candidate
            for pattern in direct_patterns:
                match = re.search(pattern, candidate_line, re.IGNORECASE)
                if match:
                    candidate = clean_report_no(match.group(1))
                    if is_plausible_report_no(candidate):
                        return candidate

    label_pattern = re.compile(r"(?i)(Report\s*No\.?|报告编号|報告編號|报告号|報告號|報告號碼|报告号码|號碼|号码|Number)")
    for idx, line in enumerate(lines[:900]):
        if not label_pattern.search(line):
            continue
        if is_historical_report_no_context(line):
            continue
        window = lines[idx + 1 : idx + 5]
        for nearby in window:
            if is_historical_report_no_context(nearby):
                continue
            candidate = extract_report_no_candidate(nearby)
            if candidate:
                return candidate
    return ""


def is_historical_report_no_context(line: str) -> bool:
    low = line.lower()
    historical_terms = [
        "original report",
        "displaces the original",
        "replace",
        "replaces",
        "invalid",
        "作废",
        "替换原报告",
        "原报告",
        "基礎上修改",
        "基础上修改",
    ]
    return any(term in low for term in historical_terms)


def clean_report_no(value: str) -> str:
    clean = normalize_spaces(value).strip(" .,:：")
    clean = repair_spaced_report_line(clean)
    clean = re.split(r"\s+(?:日期|Date)\b", clean, maxsplit=1, flags=re.IGNORECASE)[0]
    clean = re.sub(r"(?:日期|Date)\s*[:：]?.*$", "", clean, flags=re.IGNORECASE).strip(" .,:：")
    if "(" not in clean:
        clean = re.sub(r"(?<=[A-Z0-9])\)+$", "", clean)
    clean = re.sub(r"\s*/\s*", "/", clean)
    clean = re.sub(r"\s*-\s*", "-", clean)
    clean = re.sub(r"(?<=[A-Z])\s+(?=[A-Z0-9])", "", clean)
    clean = re.sub(r"(?<=[0-9])\s+(?=[0-9A-Z])", "", clean)
    clean = re.sub(r"(?<=/)\s+", "", clean)
    clean = re.sub(r"\s+(?=/)", "", clean)
    return clean


def repair_spaced_report_line(line: str) -> str:
    repaired = re.sub(r"\bN\s*o\s*\.", "No.", line, flags=re.IGNORECASE)
    repaired = re.sub(r"\bK\s*A\b", "KA", repaired)
    repaired = re.sub(r"\bR\s*o\s*H\s*S\b", "RoHS", repaired, flags=re.IGNORECASE)
    return repaired


def extract_report_no_candidate(line: str) -> str:
    for match in re.finditer(r"[A-Z0-9][A-Z0-9_\-()/\.]{4,}", line, re.IGNORECASE):
        candidate = clean_report_no(match.group(0))
        if is_plausible_report_no(candidate):
            return candidate
    return ""


def report_no_from_file_stem(file_stem: str) -> str:
    stem = re.sub(r"[_-]20\d{6}$", "", file_stem)
    stem = re.sub(r"(?i)^TestReport[_-]\d+[_-]", "", stem)
    stem = re.sub(r"(?i)^TestReport[_-]", "", stem)
    candidate = clean_report_no(stem)
    return candidate if is_plausible_report_no(candidate) else stem


def is_plausible_report_no(candidate: str) -> bool:
    if not candidate:
        return False
    low = candidate.lower()
    if low in {"report", "test", "page", "number"}:
        return False
    if any(term in low for term in ["brom", "phthalate", "report", "sample", "address", "building"]):
        return False
    if re.fullmatch(r"\d{2,7}-\d{2}-\d", candidate):
        return False
    if re.fullmatch(r"\d+", candidate):
        return len(candidate) >= 8
    if re.fullmatch(r"[()0-9\-/_.]+", candidate):
        return bool(re.search(r"\d", candidate) and len(candidate) >= 8)
    return bool(re.search(r"[A-Z]", candidate, re.I) and re.search(r"\d", candidate) and len(candidate) >= 6)


def extract_report_date(text: str, file_stem: str) -> str:
    value = first_pattern_match(
        [
            r"(?:日期\s*\(\s*Date\s*\)|Date\s*\)|日期)\s*[:：]?\s*(\d{1,2}\s*[-/]\s*[A-Z][a-z]{2,8}\.?\s*[-/]\s*\d{4})",
            r"(?:日期\s*\(\s*Date\s*\)|Date\s*\)|日期)\s*[:：]?\s*([A-Z][a-z]{2,8}\.?\s+\d{1,2},\s+\d{4})",
            r"(?:日期\s*\(\s*Date\s*\)|Date\s*\)|日期)\s*[:：]?\s*(\d{4}[./-]\d{1,2}[./-]\d{1,2})",
        ],
        text,
    )
    if value:
        normalized = normalize_report_date(value)
        if normalized:
            return normalized
    value = first_pattern_match(
        [
            r"^\s*日\s*期\s*[:：]?\s*(\d{4}[./-]\d{1,2}[./-]\d{1,2})",
            r"(?:Date|Issue\s*Date|报告日期|報告日期)\s*[:：]?\s*([A-Z][a-z]{2,8}\.?\s+\d{1,2},\s+\d{4})",
            r"(?:Date|Issue\s*Date|报告日期|報告日期)\s*[:：]?\s*(\d{1,2}\s*[-/]\s*[A-Z][a-z]{2,8}\.?\s*[-/]\s*\d{4})",
            r"(?:Date|Issue\s*Date|报告日期|報告日期)\s*[:：]?\s*(\d{4}[./-]\d{1,2}[./-]\d{1,2})",
            r"(?:Date|Issue\s*Date|报告日期|報告日期)\s*[:：]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{4})",
        ],
        text,
    )
    if value:
        normalized = normalize_report_date(value)
        if normalized:
            return normalized
    value = first_pattern_match(
        [
            r"(?:Testing\s*Period|Test\s*Period|样品检测日期|樣品檢測日期|检测日期|檢測日期)\s*[:：]?\s*(?:\d{4}[./-]\d{1,2}[./-]\d{1,2}|[A-Z][a-z]{2,8}\.?\s+\d{1,2},\s+\d{4}|\d{1,2}\s*[-/]\s*[A-Z][a-z]{2,8}\.?\s*[-/]\s*\d{4})\s*(?:-|~|至|to)\s*(\d{4}[./-]\d{1,2}[./-]\d{1,2}|[A-Z][a-z]{2,8}\.?\s+\d{1,2},\s+\d{4}|\d{1,2}\s*[-/]\s*[A-Z][a-z]{2,8}\.?\s*[-/]\s*\d{4})",
            r"(?:样品接收日期|樣品接收日期|Sample\s*Receiving\s*Date)\s*[:：]?\s*(\d{4}[./-]\d{1,2}[./-]\d{1,2})",
        ],
        text,
    )
    if value:
        normalized = normalize_report_date(value)
        if normalized:
            return normalized
    return ""


def date_from_file_stem(file_stem: str) -> str:
    match = re.search(r"(20\d{6})$", file_stem)
    if not match:
        match = re.search(r"(20\d{6})(?!\d)", file_stem)
    if match:
        raw = match.group(1)
        return normalize_report_date(f"{raw[:4]}-{raw[4:6]}-{raw[6:]}")
    return ""


def normalize_report_date(value: str) -> str:
    clean = normalize_spaces(value)
    clean = clean.replace(".", "-")
    clean = re.sub(r"\s*-\s*", "-", clean)

    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m-%d-%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%b %d, %Y",
        "%B %d, %Y",
        "%b. %d, %Y",
        "%B. %d, %Y",
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%d-%b.-%Y",
        "%d-%B.-%Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(clean, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass

    match = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", clean)
    if match:
        year, month, day = match.groups()
        try:
            return datetime(int(year), int(month), int(day)).strftime("%Y-%m-%d")
        except ValueError:
            return ""
    return ""


def extract_sample_name(text: str) -> str:
    value = first_line_match(
        [
            r"(?:Sample\s*Description|Sample\s*Name|样品名称|樣品名稱|样品描述|樣品描述)\s*[:：]?\s*(.{2,120})",
            r"(?:Description\s+of\s+Sample)\s*[:：]?\s*(.{2,120})",
            r"(?:Product\s*Name|品名)\s*[:：]?\s*(.{2,120})",
        ],
        text,
    )
    if value and "assigned by laboratory" not in value.lower():
        return value

    lines = [normalize_spaces(line) for line in text.splitlines() if normalize_spaces(line)]
    for idx, line in enumerate(lines):
        if re.fullmatch(r"(?i)SAMPLE\s+DESCRIPTION", line):
            for nearby in lines[idx + 1 : idx + 8]:
                if re.search(r"(?i)assigned by laboratory|color|style|model|p\.?\s*o\.?\s*no|country|buyer|manufacturer", nearby):
                    continue
                if ":" in nearby:
                    continue
                if len(nearby) >= 5:
                    return nearby[:120]
        if re.search(r"(?i)SAMPLE\s+DESCRIPTION\s+ASSIGNED\s+BY\s+LABORATORY", line):
            for nearby in lines[idx + 1 : idx + 6]:
                match = re.match(r"\d+\s+(.{3,120})", nearby)
                if match:
                    return normalize_spaces(match.group(1))
    return value


def extract_test_part(text: str) -> str:
    return first_line_match(
        [
            r"(?:Tested\s*(?:Component|Part)|Testing\s*Part|Test\s*Part|检测部位|測試部位|测试部位)\s*[:：]?\s*(.{2,120})",
            r"(?:Material|材质|材質)\s*[:：]?\s*(.{2,120})",
        ],
        text,
    )


def best_result_from_line(line: str, aliases: list[str]) -> str:
    low = line.lower()
    alias_positions = [low.find(alias.lower()) for alias in aliases if alias.lower() in low]
    if not alias_positions:
        return ""
    pos = min(alias_positions)
    prefix = line[max(0, pos - 60) : pos]
    if re.search(r"(?i)\bn\.?\s*d\.?\b|not\s+detected", prefix):
        return "ND"
    tail = line[pos:]
    if ("content" in low or "含量" in line) and not re.search(r"(?i)n\.?\s*d\.?|not\s+detected|pass|<\s*\d", line):
        return ""
    structured = structured_row_result(tail)
    if structured:
        return structured
    tokens = [normalize_spaces(m.group(0)) for m in RESULT_TOKEN.finditer(tail)]
    cleaned = []
    for token in tokens:
        low_token = token.lower().replace(" ", "")
        if low_token in {"pb", "cd", "hg", "pbb", "pbbs", "pbde", "pbdes", "dbp", "bbp", "dehp", "dibp"}:
            continue
        if re.fullmatch(r"\d{4}(?:\.\d+)?", low_token) and low_token.startswith(("19", "20")):
            continue
        if low_token in {"62321", "2011", "2013", "2015", "2017", "2021", "2023", "863", "65"}:
            continue
        cleaned.append(token)
    for token in cleaned:
        if re.search(r"(?i)n\.?\s*d\.?|not\s+detected|pass|negative", token):
            return token
    for token in cleaned:
        if token.strip().startswith("<"):
            return token
    return cleaned[0] if cleaned else ""


def structured_row_result(tail: str) -> str:
    clean = re.sub(r"CAS\s*#?:?\s*\d{2,7}-\d{2}-\d", " ", tail, flags=re.I)
    raw_tokens = [normalize_spaces(m.group(0)) for m in ROW_VALUE_TOKEN.finditer(clean)]
    tokens = []
    for token in raw_tokens:
        low = token.lower().replace(" ", "")
        if low in {"2011", "2013", "2015", "2017", "2021", "2023", "62321", "863", "65"}:
            continue
        tokens.append(token)

    # CTI-style rows can be "Item Result Unit MDL Unit Limit Unit",
    # e.g. "Lead (Pb) 19 mg/kg 2 mg/kg 1000 mg/kg". The first value is
    # the measured result; later values are MDL/limit and must not win.
    unit_pattern = r"(?:mg/kg|ppm|%|ug/g|µg/g|ug/cm2|µg/cm²|μg/cm2)"
    units = re.findall(unit_pattern, clean, flags=re.I)
    first_value_with_unit = re.search(rf"(?i)(<\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?)\s*{unit_pattern}", clean)
    if len(units) >= 2 and first_value_with_unit:
        return normalize_spaces(first_value_with_unit.group(1))

    after_unit = re.search(
        rf"(?i)\b{unit_pattern}\b"
        r"\s+(?:-|\d+(?:\.\d+)?)"
        r"\s+(n\.?\s*d\.?|not\s+detected|pass|negative|<\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?)\b",
        clean,
    )
    if after_unit:
        return normalize_spaces(after_unit.group(1))

    for token in tokens:
        if re.search(r"(?i)n\.?\s*d\.?|not\s+detected|pass|negative", token):
            return token

    # Common rows are Item MDL Result Limit Unit, e.g. "Pb 2 68 1000 mg/kg".
    if len(tokens) >= 3 and re.search(r"(?i)\b(mg/kg|ppm|ug/g|µg/g|%)\b", clean):
        return tokens[-2]
    return ""


def next_result_line(lines: list[str], start_idx: int) -> str:
    for line in lines[start_idx + 1 : start_idx + 5]:
        if is_non_result_context(line):
            continue
        if not re.search(r"(?i)(mg/kg|ppm|%|ug/g|繕g/g|µg/cm²|μg/cm2|nd|n\.d\.)", line):
            continue
        tokens = re.findall(r"(?i)n\.?\s*d\.?|b\.?l\.?|pass|<\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?", line)
        if len(tokens) >= 2:
            return line
    return ""


def nearby_detected_result(lines: list[str], idx: int) -> tuple[str, str]:
    window_lines = lines[idx : min(len(lines), idx + 14)]
    prev_lines = lines[max(0, idx - 2) : idx]
    for line in window_lines + list(reversed(prev_lines)):
        if "cas#:" in line.lower() or re.fullmatch(r"(?i)cas\s*#?:?.*", line.strip()):
            continue
        match = re.search(r"(?i)(?:ppm|mg/kg|ug/g|µg/g|%)?\s*(n\.?\s*d\.?|not\s+detected|pass|negative|<\s*\d+(?:\.\d+)?)\s*(?:\d+(?:\.\d+)?)?", line)
        if match and not any(skip in line.lower() for skip in ["directive", "test method", "limit", "限值", "screening"]):
            return normalize_spaces(match.group(1)), " | ".join(window_lines)[:300]
    return "", ""


def apply_adjacent_table_results(lines: list[str], values: dict[str, str]) -> dict[str, str]:
    for idx, line in enumerate(lines):
        compact = re.sub(r"[^A-Za-z0-9]+", " ", line).strip().lower()
        if re.search(r"\bpb\s+cd\s+hg\s+cr\s+br\b", compact):
            result_line = next_result_line(lines, idx)
            tokens = re.findall(r"(?i)n\.?\s*d\.?|b\.?l\.?|pass|<\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?", result_line)
            if len(tokens) >= 4:
                values["Lead"] = normalize_spaces(tokens[0])
                values["Cadmium"] = normalize_spaces(tokens[1])
                values["Mercury"] = normalize_spaces(tokens[2])
                values["Hexavalent Chromium"] = normalize_spaces(tokens[3])
                values["Lead Evidence"] = result_line[:300]
                values["Cadmium Evidence"] = result_line[:300]
                values["Mercury Evidence"] = result_line[:300]
                values["Hexavalent Chromium Evidence"] = result_line[:300]
                if len(tokens) >= 5:
                    values["PBBs"] = normalize_spaces(tokens[4])
                    values["PBDEs"] = normalize_spaces(tokens[4])
                    values["PBBs Evidence"] = result_line[:300]
                    values["PBDEs Evidence"] = result_line[:300]
        if re.search(r"dehp\s+bbp\s+dbp\s+dibp", compact):
            result_line = next_result_line(lines, idx)
            if not result_line:
                continue
            tokens = re.findall(r"(?i)n\.?\s*d\.?|pass|<\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?", result_line)
            if len(tokens) >= 4:
                for name, token in zip(["DEHP", "BBP", "DBP", "DIBP"], tokens[:4]):
                    values[name] = normalize_spaces(token)
                    values[f"{name} Evidence"] = result_line[:300]
    return values


COMPACT_RESULT_PATTERNS = {
    "Lead": [r"Lead\s*\(Pb\)", r"铅\s*\(Pb\)", r"鉛\s*/\s*Lead", r"鉛\s*\(Pb\)"],
    "Mercury": [r"Mercury\s*\(Hg\)", r"汞\s*\(Hg\)"],
    "Cadmium": [r"Cadmium\s*\(Cd\)", r"镉\s*\(Cd\)", r"鎘\s*/\s*Cadmium", r"鎘\s*\(Cd\)"],
    "Hexavalent Chromium": [r"Hexavalent\s+Chromium\s*\(Cr\s*\(?VI\)?\)", r"六价铬\s*\(Cr\s*\(?VI\)?\)", r"六價鉻\s*/\s*Hexavalent\s+Chromium", r"Chromium\s*\(VI\)"],
    "PBBs": [r"Sum\s+of\s+PBBs", r"多溴联苯之和\s*\(PBBs\)", r"多溴聯苯總和", r"Polybrominated\s+Biphenyls\s*\(PBBs\)"],
    "PBDEs": [r"Sum\s+of\s+PBDEs", r"多溴二苯醚之和\s*\(PBDEs\)", r"多溴聯苯醚總和", r"Polybrominated\s+Diphenyl\s+Ethers\s*\(PBDEs\)"],
    "BBP": [r"Benzyl\s+Butyl\s+Phthalate\s*\(BBP\)", r"Butyl\s*Benzyl\s+Phthalate\s*\(BBP\)", r"邻苯二甲酸丁苄酯\s*\(BBP\)", r"鄰苯二甲酸丁苄酯\s*\(BBP\)", r"鄰苯二甲酸苯基丁酯"],
    "DBP": [r"Di-?butyl\s+Phthalate\s*\(DBP\)", r"Di-n-butyl\s+Phthalate\s*\(DBP\)", r"Dibutyl\s+Phthalate\s*\(DBP\)", r"邻苯二甲酸二丁酯\s*\(DBP\)", r"鄰苯二甲酸二丁酯\s*\(DBP\)"],
    "DEHP": [r"Di-?\(?2-ethyl\s*hexyl\)?\s+Phthalate\s*\(DEHP\)", r"Di\s*\(?2-ethyl\s*hexyl\)?\s+Phthalate\s*\(DEHP\)", r"Bis-?\s*\(?2-ethyl\s*hexyl\)?\s+phthalate\s*\(DEHP\)", r"邻苯二甲酸二\s*\(?2-乙基.*?\(DEHP\)", r"鄰苯二甲酸二\s*\(?2-乙基.*?\(DEHP\)", r"邻苯二甲酸二异辛酯\s*\(DEHP\)"],
    "DIBP": [r"Diisobutyl\s+Phthalates?\s*\(DIBP\)", r"Di-\(iso-butyl\)\s+phthalate\s*\(DIBP\)", r"邻苯二甲酸二异丁酯\s*\(DIBP\)", r"鄰苯二甲酸二異丁酯\s*\(DIBP\)"],
}


def compact_result_overrides(text: str) -> dict[str, str]:
    compact = normalize_compact_text(text)
    overrides: dict[str, str] = {}
    for name, patterns in COMPACT_RESULT_PATTERNS.items():
        for pattern in patterns:
            result = compact_result_after_pattern(compact, pattern)
            if result:
                value, evidence = result
                overrides[name] = value
                overrides[f"{name} Evidence"] = evidence
                break
    return overrides


def normalize_compact_text(text: str) -> str:
    compact = normalize_spaces(text)
    compact = compact.replace("\u00a0", " ")
    compact = re.sub(r"(?<=\w)\.(?=\w)", ". ", compact)
    compact = re.sub(r"\s+", " ", compact)
    return compact


def compact_result_after_pattern(compact: str, pattern: str) -> tuple[str, str] | None:
    regex = re.compile(
        pattern
        + r".{0,100}?(?:-|\d[\d,]*)?\s*(?:mg/kg)?\s*\.?\s*(?:-|\d+(?:\.\d+)?)?\s*(N\.?\s*D\.?|ND|Not\s+Detected|<\s*\d+(?:\.\d+)?)",
        re.IGNORECASE,
    )
    match = regex.search(compact)
    if match:
        value = normalize_result_token(match.group(1))
        return value, match.group(0)[:300]
    return None


def normalize_result_token(value: str) -> str:
    clean = normalize_spaces(value)
    clean = re.sub(r"(?i)^n\.?\s*d\.?$", "ND", clean)
    return clean


def extract_substances(text: str) -> dict[str, str]:
    lines = [normalize_spaces(line) for line in text.splitlines() if normalize_spaces(line)]
    values = {name: "" for name, _aliases in SUBSTANCES}
    evidence = {name: "" for name, _aliases in SUBSTANCES}

    values = apply_adjacent_table_results(lines, values)
    apply_sgs_taiwan_split_table_results(text, values)
    compact_overrides = compact_result_overrides(text)

    for name, aliases in SUBSTANCES:
        if values.get(name):
            continue
        for line in lines:
            low_line = line.lower()
            if is_non_result_context(line):
                continue
            if any(skip in low_line for skip in ["directive", "2011/65", "2015/863", "test method", "检测方法", "檢測方法", "limit", "限值", "screening limit", "mdl", "method detection", "cas#:"]):
                continue
            if "%" in line and not re.search(r"(?i)n\.?\s*d\.?|not\s+detected|pass|<\s*\d", line):
                continue
            if name in {"BBP", "DBP", "DEHP", "DIBP"} and (
                "phthalates (" in low_line or "邻苯二甲酸酯" in line or "鄰苯二甲酸酯" in line
            ):
                continue
            if not re.search(r"(?i)n\.?\s*d\.?|not\s+detected|pass|<\s*\d|\d+(?:\.\d+)?\s*(?:mg/kg|ppm|%)", line):
                continue
            if not any(re.search(rf"(?<![a-z0-9]){re.escape(alias)}(?![a-z0-9])", line, re.I) for alias in aliases):
                continue
            result = best_result_from_line(line, aliases)
            if result:
                values[name] = result
                evidence[name] = line[:300]
                break
        if not values[name]:
            for idx, line in enumerate(lines):
                low_line = line.lower()
                if is_non_result_context(line):
                    continue
                if any(skip in low_line for skip in ["directive", "2011/65", "2015/863", "test method", "检测方法", "檢測方法", "limit", "限值", "screening limit", "mdl", "method detection", "cas#:"]):
                    continue
                if "%" in line and not re.search(r"(?i)n\.?\s*d\.?|not\s+detected|pass|<\s*\d", line):
                    continue
                if name in {"BBP", "DBP", "DEHP", "DIBP"} and (
                    "phthalates (" in low_line or "邻苯二甲酸酯" in line or "鄰苯二甲酸酯" in line
                ):
                    continue
                if any(alias_in_text(alias, line) for alias in aliases):
                    window = " ".join(lines[idx : idx + 3])
                    if not re.search(r"(?i)n\.?\s*d\.?|not\s+detected|pass|<\s*\d|\d+(?:\.\d+)?\s*(?:mg/kg|ppm|%)", window):
                        continue
                    result = best_result_from_line(window, aliases)
                    if result:
                        values[name] = result
                        evidence[name] = window[:300]
                        break
                    nearby_value, nearby_evidence = nearby_detected_result(lines, idx)
                    if nearby_value:
                        values[name] = nearby_value
                        evidence[name] = nearby_evidence
                        break

    apply_mentioned_nd_fallback(text, values)
    apply_sgs_taiwan_rohs10_nd_fallback(text, values)

    for name, _aliases in SUBSTANCES:
        if compact_overrides.get(name) and not values.get(name):
            values[name] = compact_overrides[name]
            values[f"{name} Evidence"] = compact_overrides.get(f"{name} Evidence", "")
        elif evidence[name]:
            values[f"{name} Evidence"] = evidence[name]
    return values


def apply_sgs_taiwan_split_table_results(text: str, values: dict[str, str]) -> None:
    compact = normalize_compact_text(text)
    low = compact.lower()
    squashed = squash_for_keyword(compact)
    if "sgs taiwan" not in low and "sgstaiwan" not in squashed and "台灣檢驗科技" not in compact:
        return
    if "測試項目" not in compact and "test items" not in low:
        return

    lines = [normalize_spaces(line) for line in text.splitlines() if normalize_spaces(line)]
    result_row = re.compile(
        r"(?i)^(mg/kg|ppm|%)\s+(\d+(?:\.\d+)?|-)\s+"
        r"(n\.?\s*d\.?|nd|not\s+detected|<\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?)\s+"
        r"(\d+(?:\.\d+)?|-)\b"
    )
    start_idx = 0
    for idx, line in enumerate(lines):
        if re.search(r"(?i)測試方法\s+單位\s+MDL\s+限值|Method\)\s*\(Unit\)\s*\(Limit\)", line):
            start_idx = idx
            break

    rows: list[tuple[str, str, str, str, str]] = []
    for line in lines[start_idx : start_idx + 120]:
        match = result_row.search(line)
        if match:
            unit, mdl, result, limit = match.groups()
            rows.append((unit, mdl, normalize_result_token(result), limit, line))
        if len(rows) >= 4:
            break

    if len(rows) < 4:
        return

    sequence = ["Cadmium", "Lead", "Mercury", "Hexavalent Chromium"]
    for name, (unit, mdl, result, limit, line) in zip(sequence, rows):
        if values.get(name):
            continue
        values[name] = result
        values[f"{name} Evidence"] = (
            "SGS Taiwan split result table: "
            f"{line}; mapped by row order {', '.join(sequence)}."
        )


def apply_sgs_taiwan_rohs10_nd_fallback(text: str, values: dict[str, str]) -> None:
    compact = normalize_compact_text(text)
    low = compact.lower()
    required_terms = [
        "cadmium",
        "lead",
        "mercury",
        "cr(vi)",
        "pbb",
        "pbde",
        "dbp",
        "bbp",
        "dehp",
        "dibp",
    ]
    squashed = squash_for_keyword(compact)
    if "sgs taiwan" not in low and "sgstaiwan" not in squashed and "台灣檢驗科技" not in compact:
        return
    if "rohs directive" not in low and "2015/863" not in low:
        return
    if "comply with the limits" not in low and "pass" not in low:
        return
    if not all(term in low for term in required_terms):
        return
    if len(re.findall(r"(?i)\b(?:mg/kg|ppm|%)\s+(?:\d+(?:\.\d+)?|-)\s+n\.?\s*d\.?\b", compact)) < 8:
        return

    for name, _aliases in SUBSTANCES:
        if values.get(name):
            continue
        values[name] = "n.d."
        values[f"{name} Evidence"] = (
            "Fallback: SGS Taiwan RoHS summary lists all 10 restricted substances as compliant, "
            "and the extracted result table contains repeated n.d. rows."
        )


def apply_mentioned_nd_fallback(text: str, values: dict[str, str]) -> None:
    compact = normalize_compact_text(text)
    if not re.search(r"(?i)n\.?\s*d\.?", compact):
        return
    lines = [normalize_spaces(line) for line in text.splitlines() if normalize_spaces(line)]
    for name, aliases in SUBSTANCES:
        if values.get(name):
            continue
        if name in {"Lead", "Mercury", "Cadmium", "Hexavalent Chromium"}:
            continue
        for idx, line in enumerate(lines):
            if is_non_result_context(line):
                continue
            if not any(alias_in_text(alias, line) for alias in aliases):
                continue
            window = " ".join(lines[idx : idx + 3])
            if is_non_result_context(window):
                continue
            if re.search(r"(?i)\bn\.?\s*d\.?\b", window) and re.search(
                r"(?i)(mg/kg|ppm|%|cas\s*no|test\s*item|result)", window
            ):
                values[name] = "n.d."
                values[f"{name} Evidence"] = "Fallback: substance row contains n.d. in a result-table context."
                break
    apply_rohs_pass_nd_table_fallback(compact, values)


def alias_in_text(alias: str, text: str) -> bool:
    if len(alias) <= 4 and alias.isascii() and alias.isalnum():
        return bool(re.search(rf"(?<![a-z0-9]){re.escape(alias)}(?![a-z0-9])", text, re.I))
    return alias.lower() in text.lower()


def is_non_result_context(text: str) -> bool:
    low = text.lower()
    skip_terms = [
        "sample name",
        "sample description",
        "product name",
        "model",
        "rohs＆pb free",
        "rohs&pb free",
        "pb free",
        "test requested",
        "test requirement conclusion",
        "directive",
        "2011/65",
        "2015/863",
        "flow chart",
        "remark",
        "note",
        "content pass",
        "test conclusion",
        "summary of test",
        "comment -- pass",
    ]
    return any(term in low for term in skip_terms)


def apply_rohs_pass_nd_table_fallback(compact: str, values: dict[str, str]) -> None:
    low = compact.lower()
    if "rohs" not in low or "pass" not in low or not re.search(r"(?i)\bn\.?\s*d\.?\b", compact):
        return
    if "phthalate" not in low:
        return
    if not (values.get("DBP") and values.get("DIBP")):
        return
    for name in ["BBP", "DEHP"]:
        if not values.get(name):
            values[name] = "ND"
            values[f"{name} Evidence"] = "Fallback: RoHS phthalate table OCR shows PASS/ND and adjacent DBP/DIBP rows were extracted."


def clear_generated_assessment_notes(record: PdfRecord) -> None:
    generated_markers = [
        "No RoHS directive",
        "不為RoHS",
        "報告號碼未能",
        "報告日期未能",
        "無法判讀RoHS",
        "RoHS相關性信心不足",
        "No usable text was extracted",
    ]
    record.notes = [
        note for note in record.notes if not any(marker in note for marker in generated_markers)
    ]


def populate_record_from_current_text(record: PdfRecord) -> tuple[str, str]:
    clear_generated_assessment_notes(record)
    text = "\n".join(page.text for page in record.pages)
    compact = normalize_spaces(text)
    file_stem = Path(record.file_name).stem
    record.lab_name = ""
    record.report_no = ""
    record.report_date = ""
    record.sample_name = ""
    record.test_part = ""
    record.substances = {}
    record.is_rohs_related = None
    record.rohs_confidence = 0.0
    record.non_rohs_confidence = 0.0
    record.result_confidence = 0.0

    if compact:
        record.lab_name = detect_lab(compact)
        record.report_no = extract_report_no(text, file_stem)
        record.report_date = extract_report_date(text, file_stem)
        record.sample_name = extract_sample_name(text)
        record.test_part = extract_test_part(text)
        (
            record.is_rohs_related,
            record.rohs_confidence,
            record.non_rohs_confidence,
        ) = assess_rohs_relevance(text)
        if record.is_rohs_related is True:
            if record.extraction_status in {"不為RoHS", "無法判讀", "low_or_no_text", "text_failed_or_encrypted"}:
                record.extraction_status = "text_extracted"
            record.substances = extract_substances(text)
        elif record.is_rohs_related is False:
            record.extraction_status = "不為RoHS"
            record.notes.append(
                f"不為RoHS：非RoHS信心={record.non_rohs_confidence:.2f}，RoHS信心={record.rohs_confidence:.2f}。"
            )
        else:
            record.extraction_status = "無法判讀"
            record.notes.append(
                f"RoHS相關性信心不足：RoHS信心={record.rohs_confidence:.2f}，非RoHS信心={record.non_rohs_confidence:.2f}，需人工確認。"
            )
    else:
        record.extraction_status = "無法判讀"

    if not record.report_no:
        record.report_no = report_no_from_file_stem(file_stem)
        record.notes.append("報告號碼未能從PDF內容判讀，已由檔名補入，需人工確認。")
    if not record.report_date:
        file_date = date_from_file_stem(file_stem)
        if file_date:
            record.report_date = file_date
            record.notes.append("報告日期未能從PDF內容判讀，已由檔名補入，需人工確認。")
    if compact and not record.substances and record.extraction_status != "不為RoHS":
        record.notes.append("無法判讀RoHS結果：未能穩定抽取限制物質結果，需人工確認。")
    refresh_record_confidence(record)

    if record_has_ocr_text(record):
        suspicious_numeric_names = []
        for name, _aliases in SUBSTANCES:
            value = record.substances.get(name, "")
            if not value or re.search(r"(?i)n\.?\s*d\.?|not\s+detected|pass|negative", value):
                continue
            evidence = record.substances.get(f"{name} Evidence", "")
            if re.search(r"(?i)(?:limit|mdl).*(?:\d|mg/kg|ppm)|\d+\.\d+", evidence):
                continue
            if re.search(r"\d", value):
                suspicious_numeric_names.append(name)
        if suspicious_numeric_names:
            record.notes.append(
                "OCR numeric anomaly detected; targeted source-page review is required for: "
                + ", ".join(suspicious_numeric_names)
            )

    if record.needs_ocr:
        record.notes.append("Low/no extractable text. OCR or another PDF engine is needed.")
    if record.needs_pdf_fallback:
        record.notes.append("PDF fallback is needed, possibly due to AES encryption or unsupported PDF objects.")
    if not compact:
        record.notes.append("No usable text was extracted.")
    return text, compact


def should_add_ocr_after_text_parse(record: PdfRecord, text: str, compact: str) -> str:
    if record_has_ocr_text(record):
        return ""
    if record.needs_ocr or record.extraction_status in {"low_or_no_text", "text_failed_or_encrypted"}:
        return "PDF text layer is low, missing, or failed."
    if is_probably_garbled_text(text):
        return "PDF text layer appears garbled."

    substance_count = sum(bool(record.substances.get(name)) for name, _ in SUBSTANCES)
    low = compact.lower()
    has_rohs10_summary = all(
        term in low
        for term in ["lead", "cadmium", "mercury", "pbb", "pbde", "dbp", "bbp", "dehp", "dibp"]
    ) and ("rohs" in low or "2015/863" in low)
    if has_rohs10_summary and substance_count < 10:
        return "RoHS 10-item summary was found but fewer than 10 substance results were extracted."
    if record.is_rohs_related is True and record.result_confidence < 0.75:
        return "RoHS report confidence is below 0.75."
    if record.is_rohs_related is None and record.rohs_confidence >= 0.35:
        return "RoHS/non-RoHS relevance confidence is inconclusive."
    return ""


def should_add_markitdown_after_text_parse(record: PdfRecord, text: str, compact: str) -> str:
    if record_has_markitdown_text(record):
        return ""
    if not compact:
        return ""
    if record.is_rohs_related is False:
        return ""

    substance_count = sum(bool(record.substances.get(name)) for name, _ in SUBSTANCES)
    low = compact.lower()
    has_rohs10_summary = all(
        term in low
        for term in ["lead", "cadmium", "mercury", "pbb", "pbde", "dbp", "bbp", "dehp", "dibp"]
    ) and ("rohs" in low or "2015/863" in low)

    if record.is_rohs_related is True:
        if not record.report_no:
            return "RoHS report number is missing after PDF text parsing."
        if not record.report_date:
            return "RoHS report date is missing after PDF text parsing."
        if has_rohs10_summary and substance_count < 10:
            return "RoHS 10-item summary was found but fewer than 10 substance results were extracted."
        if record.result_confidence < 0.75:
            return "RoHS report confidence is below 0.75."
    if record.is_rohs_related is None and record.rohs_confidence >= 0.35:
        return "RoHS/non-RoHS relevance confidence is inconclusive."
    return ""


def page_bundles(record: PdfRecord) -> list[tuple[int, list[PageText]]]:
    bundled: dict[int, list[PageText]] = {}
    for page in record.pages:
        bundled.setdefault(page.page, []).append(page)
    return [(page_no, bundled[page_no]) for page_no in sorted(bundled)]


def report_no_from_page_bundle(bundle: list[PageText]) -> str:
    text = "\n".join(page.text for page in bundle if page.text)
    if not text:
        return ""
    return extract_report_no(text, "")


def clone_record_for_page_group(
    source: PdfRecord,
    pages: list[PageText],
    segment_index: int,
    segment_count: int,
    hinted_report_no: str,
) -> PdfRecord:
    page_numbers = sorted({page.page for page in pages})
    clone = PdfRecord(
        file_name=source.file_name,
        file_path=source.file_path,
        file_size=source.file_size,
        page_count=len(page_numbers),
        encrypted=source.encrypted,
        text_char_count=sum(page.char_count for page in pages),
        extraction_status=source.extraction_status,
        needs_ocr=False,
        needs_pdf_fallback=source.needs_pdf_fallback,
        pages=pages,
        segment_id=f"{segment_index}/{segment_count}",
    )
    if clone.text_char_count < 500:
        clone.extraction_status = "low_or_no_text"
        clone.needs_ocr = True
    page_label = ",".join(str(page_no) for page_no in page_numbers)
    clone.notes.append(f"一檔多報告模式：第 {segment_index}/{segment_count} 段，來源頁碼 {page_label}。")
    if hinted_report_no:
        clone.notes.append(f"一檔多報告模式：頁面偵測報告號碼 {hinted_report_no}。")
    return clone


def split_record_by_report_number(record: PdfRecord) -> list[PdfRecord]:
    bundles = page_bundles(record)
    if not bundles:
        return [record]

    raw_groups: list[tuple[str, list[PageText]]] = []
    current_report_no = ""
    current_pages: list[PageText] = []

    for _page_no, bundle in bundles:
        detected_report_no = report_no_from_page_bundle(bundle)
        starts_new_report = (
            bool(detected_report_no)
            and bool(current_pages)
            and bool(current_report_no)
            and detected_report_no != current_report_no
        )
        if starts_new_report:
            raw_groups.append((current_report_no, current_pages))
            current_pages = []
        if detected_report_no and not current_report_no:
            current_report_no = detected_report_no
        elif starts_new_report:
            current_report_no = detected_report_no
        current_pages.extend(bundle)

    if current_pages:
        raw_groups.append((current_report_no, current_pages))

    if len(raw_groups) <= 1:
        record.segment_id = "1/1"
        record.notes.append("一檔多報告模式：未偵測到多個報告號碼，依一般單份報告處理。")
        return [record]

    return [
        clone_record_for_page_group(record, pages, idx, len(raw_groups), report_no)
        for idx, (report_no, pages) in enumerate(raw_groups, start=1)
    ]


def enrich_record(record: PdfRecord, ai_config: AiConfig | None = None) -> PdfRecord:
    text, compact = populate_record_from_current_text(record)
    markitdown_reason = should_add_markitdown_after_text_parse(record, text, compact)
    if markitdown_reason and add_markitdown_text(record, markitdown_reason):
        text, compact = populate_record_from_current_text(record)
    reason = should_add_ocr_after_text_parse(record, text, compact)
    if reason and add_ocr_text(record, reason):
        text, compact = populate_record_from_current_text(record)
    if ai_config:
        apply_ai_assistance(record, ai_config)
    return record


def write_json(record: PdfRecord, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(record.file_name).stem)
    if record.segment_id:
        safe_segment = re.sub(r"[^A-Za-z0-9_.-]+", "_", record.segment_id)
        safe_name = f"{safe_name}__segment_{safe_segment}"
    if record.report_no:
        safe_report_no = re.sub(r"[^A-Za-z0-9_.-]+", "_", record.report_no)
        if safe_report_no and safe_report_no not in safe_name:
            safe_name = f"{safe_name}__{safe_report_no}"
    path = output_dir / f"{safe_name}.json"
    with path.open("w", encoding="utf-8") as fh:
        json.dump(asdict(record), fh, ensure_ascii=False, indent=2)


UNIT_PATTERN = re.compile(
    r"(?i)(mg/kg|ppm|%|ug/g|µg/g|μg/g|ug/cm2|µg/cm2|μg/cm2|ug/cm²|µg/cm²|μg/cm²)"
)


def split_value_and_unit(value: str, evidence: str = "") -> tuple[str, str]:
    raw_value = normalize_spaces(str(value or ""))
    raw_value = re.sub(r"(?i)\bN\s+D\b", "ND", raw_value)
    if not raw_value:
        return "", ""

    unit = ""
    match = UNIT_PATTERN.search(raw_value)
    if match:
        unit = normalize_unit(match.group(1))
        raw_value = normalize_spaces(UNIT_PATTERN.sub("", raw_value))
    else:
        evidence_match = UNIT_PATTERN.search(evidence or "")
        if evidence_match:
            unit = normalize_unit(evidence_match.group(1))

    return raw_value, unit


def normalize_unit(unit: str) -> str:
    value = unit.replace("μ", "µ")
    value = value.replace("cm2", "cm²")
    lower = value.lower()
    if lower == "ppm":
        return "ppm"
    if lower == "mg/kg":
        return "mg/kg"
    if lower in {"ug/g", "µg/g"}:
        return "µg/g"
    if lower in {"ug/cm²", "µg/cm²"}:
        return "µg/cm²"
    if value == "%":
        return "%"
    return value


def result_cells(record: PdfRecord, substance_name: str) -> list[str]:
    value = record.substances.get(substance_name, "")
    evidence = record.substances.get(f"{substance_name} Evidence", "")
    split_value, unit = split_value_and_unit(value, evidence)
    if split_value and not unit and substance_name != "Hexavalent Chromium":
        unit = "mg/kg"
    return [split_value, unit]


def result_cells_with_confidence(record: PdfRecord, substance_name: str) -> list[str | float]:
    value, unit = result_cells(record, substance_name)
    confidence = round(substance_confidence(record, substance_name), 2)
    return [value, unit, confidence]


def rohs_decision_label(record: PdfRecord) -> str:
    if record.is_rohs_related is True:
        return "RoHS"
    if record.is_rohs_related is False:
        return "非RoHS"
    return "無法判讀"


def build_workbook(records: list[PdfRecord], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "RoHS_Result"
    audit = wb.create_sheet("Extraction_Audit")
    inv = wb.create_sheet("PDF_Inventory")

    result_headers = [
        "報告號碼",
        "報告日期",
        "實驗室名",
        "樣品名稱",
        "測試部位",
        *[header for name, _ in SUBSTANCES for header in (name, "單位", "信心")],
        "來源檔案",
        "報告分段",
        "AI判定",
        "AI判定欄位",
        "AI模型",
        "RoHS判定",
        "RoHS信心",
        "非RoHS信心",
        "擷取信心",
        "信心分數",
        "狀態",
        "備註",
    ]
    ws.append(result_headers)
    for record in records:
        ws.append(
            [
                record.report_no,
                record.report_date,
                record.lab_name,
                record.sample_name,
                record.test_part,
                *[cell for name, _ in SUBSTANCES for cell in result_cells_with_confidence(record, name)],
                record.file_name,
                record.segment_id,
                "AI判定" if record.ai_assisted else "",
                ", ".join(record.ai_fields),
                record.ai_model,
                rohs_decision_label(record),
                record.rohs_confidence,
                record.non_rohs_confidence,
                record.result_confidence,
                record.confidence,
                record.extraction_status,
                "; ".join(dict.fromkeys(record.notes)),
            ]
        )

    audit_headers = ["來源檔案", "欄位", "值", "單位", "來源類型", "物質信心", "證據文字"]
    audit.append(audit_headers)
    for record in records:
        for name, _ in SUBSTANCES:
            value, unit = result_cells(record, name)
            audit.append(
                [
                    record.file_name,
                    name,
                    value,
                    unit,
                    substance_source(record, name),
                    round(substance_confidence(record, name), 2),
                    record.substances.get(f"{name} Evidence", ""),
                ]
            )

    inv_headers = [
        "來源檔案",
        "大小(bytes)",
        "頁數",
        "是否加密",
        "文字字數",
        "抽取狀態",
        "報告分段",
        "AI判定",
        "AI判定欄位",
        "AI模型",
        "RoHS判定",
        "RoHS信心",
        "非RoHS信心",
        "擷取信心",
        "需要OCR",
        "需要PDF fallback",
        "錯誤/備註",
    ]
    inv.append(inv_headers)
    for record in records:
        inv.append(
            [
                record.file_name,
                record.file_size,
                record.page_count,
                record.encrypted,
                record.text_char_count,
                record.extraction_status,
                record.segment_id,
                record.ai_assisted,
                ", ".join(record.ai_fields),
                record.ai_model,
                rohs_decision_label(record),
                record.rohs_confidence,
                record.non_rohs_confidence,
                record.result_confidence,
                record.needs_ocr,
                record.needs_pdf_fallback,
                "; ".join(dict.fromkeys(record.notes)),
            ]
        )

    for sheet in [ws, audit, inv]:
        style_sheet(sheet)

    wb.save(output_path)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 8
        for cell in column_cells[:80]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 48)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def find_pdfs(input_dir: Path) -> list[Path]:
    return sorted(path for path in input_dir.glob("*.pdf") if path.is_file())


def safe_print(text: str) -> None:
    encoding = sys.stdout.encoding or "utf-8"
    print(text.encode(encoding, errors="replace").decode(encoding, errors="replace"))


def print_summary(records: list[PdfRecord], output_path: Path) -> None:
    safe_print(f"Processed PDFs: {len(records)}")
    safe_print(f"Excel output: {output_path}")
    for record in records:
        flags = []
        if record.needs_ocr:
            flags.append("needs OCR")
        if record.needs_pdf_fallback:
            flags.append("needs fallback")
        flag_text = f" ({', '.join(flags)})" if flags else ""
        safe_print(
            f"- {record.file_name}: {record.extraction_status}, "
            f"chars={record.text_char_count}, confidence={record.confidence}{flag_text}"
        )


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract first-pass RoHS results from PDFs.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Folder containing PDF reports.")
    parser.add_argument(
        "--file",
        type=Path,
        action="append",
        default=[],
        help="Optional PDF file path. Can be used multiple times for targeted tests.",
    )
    parser.add_argument("--json-dir", type=Path, default=DEFAULT_JSON, help="Folder for per-PDF JSON files.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output .xlsx path.")
    parser.add_argument(
        "--mode",
        choices=["normal", "multi-report"],
        default="normal",
        help="normal: one PDF is one report; multi-report: split one PDF into multiple reports by report number.",
    )
    parser.add_argument("--no-ai", action="store_true", help="Disable AI assistance.")
    parser.add_argument("--ai-provider", choices=["gemini", "nvidia"], default="gemini", help="AI provider for assistance.")
    parser.add_argument("--ai-model", default="gemini-3.1-flash-lite", help="AI model for assistance.")
    parser.add_argument(
        "--ai-low-confidence",
        type=float,
        default=0.6,
        help="Run AI verification for RoHS reports below this confidence.",
    )
    parser.add_argument("--ai-log", type=Path, default=DEFAULT_AI_LOG, help="JSONL log for AI-assisted decisions.")
    parser.add_argument(
        "--ai-log-files",
        type=Path,
        default=DEFAULT_AI_LOG_FILES,
        help="Folder to copy RoHS low-confidence source PDFs that required AI assistance.",
    )
    parser.add_argument("--no-ai-log", action="store_true", help="Disable AI assistance logging.")
    args = parser.parse_args()

    input_dir = args.input.resolve()
    json_dir = args.json_dir.resolve()
    output_path = args.output.resolve()

    input_dir.mkdir(parents=True, exist_ok=True)
    json_dir.mkdir(parents=True, exist_ok=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    pdfs = [path.resolve() for path in args.file if path.exists()]
    if not pdfs:
        pdfs = find_pdfs(input_dir)
    if not pdfs:
        print(f"No PDF files found in: {input_dir}")
        print("Put reports in input_pdfs and run this script again.")
        build_workbook([], output_path)
        return 0

    if not shutil.which("tesseract"):
        print("Note: tesseract was not found on PATH. OCR fallback is only marked, not executed.")

    env_file_values = load_env_file(ROOT / ".env")
    ai_config = AiConfig(
        enabled=not args.no_ai,
        provider=args.ai_provider,
        api_key=env_value("GEMINI_API_KEY", env_file_values)
        or env_value("GOOGLE_API_KEY", env_file_values),
        nvidia_api_key=env_value("NVIDIA_API_KEY", env_file_values),
        model=args.ai_model,
        low_confidence_threshold=args.ai_low_confidence,
        log_path=None if args.no_ai_log else args.ai_log,
        log_files_dir=None if args.no_ai_log else args.ai_log_files,
    )

    records = []
    for pdf_path in pdfs:
        source_record = extract_text_with_pypdf(pdf_path)
        if args.mode == "multi-report":
            if source_record.needs_ocr or source_record.text_char_count < 500:
                add_ocr_text(
                    source_record,
                    "Multi-report mode scans the whole PDF for report numbers.",
                    max_pages=source_record.page_count,
                )
            candidate_records = split_record_by_report_number(source_record)
        else:
            candidate_records = [source_record]
        for record in candidate_records:
            record = enrich_record(record, ai_config)
            write_json(record, json_dir)
            records.append(record)

    build_workbook(records, output_path)
    print_summary(records, output_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
